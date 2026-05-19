#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RAW_DIR = Path("data/raw/microdados_censo_da_educacao_superior_2024/dados")
IES_CSV = RAW_DIR / "MICRODADOS_ED_SUP_IES_2024.CSV"
CURSOS_CSV = RAW_DIR / "MICRODADOS_CADASTRO_CURSOS_2024.CSV"
DEFAULT_OUTPUT = Path("lista_universidades_publicas_historia.xlsx")


ORG_MAP = {
    1: "Universidade",
    2: "Centro Universitário",
    3: "Faculdade",
    4: "Instituto Federal/Cefet",
    5: "Centro Federal de Educação Tecnológica",
}

REDE_MAP = {
    1: "Pública",
    2: "Privada",
}

CATEGORIA_MAP = {
    1: "Pública Federal",
    2: "Pública Estadual",
    3: "Pública Municipal",
    4: "Privada com fins lucrativos",
    5: "Privada sem fins lucrativos",
    7: "Especial",
}

GRAU_MAP = {
    1: "Bacharelado",
    2: "Licenciatura",
    3: "Tecnológico",
    4: "Bacharelado e Licenciatura",
}

MODALIDADE_MAP = {
    1: "Presencial",
    2: "EaD",
}

NIVEL_MAP = {
    1: "Graduação",
    2: "Sequencial de formação específica",
}


def normalize(value: object) -> str:
    text = "" if pd.isna(value) else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text).strip().upper()


def require_inputs() -> None:
    missing = [path for path in [IES_CSV, CURSOS_CSV] if not path.exists()]
    if missing:
        paths = "\n".join(f"- {path}" for path in missing)
        raise SystemExit(
            "Arquivos de entrada não encontrados. Extraia os CSVs do ZIP do Inep antes de rodar:\n"
            f"{paths}"
        )


def load_public_universities() -> pd.DataFrame:
    usecols = [
        "NU_ANO_CENSO",
        "CO_IES",
        "NO_IES",
        "SG_IES",
        "NO_REGIAO_IES",
        "NO_UF_IES",
        "SG_UF_IES",
        "NO_MUNICIPIO_IES",
        "TP_ORGANIZACAO_ACADEMICA",
        "TP_REDE",
        "TP_CATEGORIA_ADMINISTRATIVA",
        "DS_ENDERECO_IES",
        "DS_NUMERO_ENDERECO_IES",
        "DS_COMPLEMENTO_ENDERECO_IES",
        "NO_BAIRRO_IES",
        "NU_CEP_IES",
    ]
    ies = pd.read_csv(IES_CSV, sep=";", encoding="latin1", usecols=usecols, low_memory=False)
    public_universities = ies[
        (ies["TP_ORGANIZACAO_ACADEMICA"] == 1)
        & (ies["TP_REDE"] == 1)
    ].copy()

    public_universities["organizacao_academica"] = public_universities[
        "TP_ORGANIZACAO_ACADEMICA"
    ].map(ORG_MAP)
    public_universities["rede"] = public_universities["TP_REDE"].map(REDE_MAP)
    public_universities["categoria_administrativa"] = public_universities[
        "TP_CATEGORIA_ADMINISTRATIVA"
    ].map(CATEGORIA_MAP)
    public_universities["endereco_ies"] = (
        public_universities[
            ["DS_ENDERECO_IES", "DS_NUMERO_ENDERECO_IES", "DS_COMPLEMENTO_ENDERECO_IES"]
        ]
        .fillna("")
        .astype(str)
        .agg(" ".join, axis=1)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    public_universities["nome_normalizado"] = public_universities["NO_IES"].map(normalize)

    return public_universities.sort_values(["SG_UF_IES", "NO_IES"]).reset_index(drop=True)


def load_history_courses(public_ies_ids: set[int]) -> pd.DataFrame:
    usecols = [
        "TP_DIMENSAO",
        "CO_IES",
        "CO_CURSO",
        "NO_CURSO",
        "NO_CINE_ROTULO",
        "TP_GRAU_ACADEMICO",
        "TP_MODALIDADE_ENSINO",
        "TP_NIVEL_ACADEMICO",
        "QT_CURSO",
        "QT_VG_TOTAL",
        "QT_MAT",
        "NO_REGIAO",
        "NO_UF",
        "SG_UF",
        "NO_MUNICIPIO",
    ]

    chunks: list[pd.DataFrame] = []
    for chunk in pd.read_csv(
        CURSOS_CSV,
        sep=";",
        encoding="latin1",
        usecols=usecols,
        chunksize=200_000,
        low_memory=False,
    ):
        chunk = chunk[
            (chunk["TP_DIMENSAO"] == 1)
            & (chunk["CO_IES"].isin(public_ies_ids))
        ].copy()
        course_name = chunk["NO_CURSO"].map(normalize)
        cine_name = chunk["NO_CINE_ROTULO"].map(normalize)
        is_history = course_name.str.contains(r"\bHISTORIA\b", regex=True, na=False)
        is_art_history = cine_name.str.contains("HISTORIA DA ARTE", na=False)
        chunks.append(chunk[is_history & ~is_art_history])

    if not chunks:
        return pd.DataFrame(columns=usecols)

    courses = pd.concat(chunks, ignore_index=True).drop_duplicates("CO_CURSO")
    courses["grau_academico"] = courses["TP_GRAU_ACADEMICO"].map(GRAU_MAP)
    courses["modalidade"] = courses["TP_MODALIDADE_ENSINO"].map(MODALIDADE_MAP)
    courses["nivel_academico"] = courses["TP_NIVEL_ACADEMICO"].map(NIVEL_MAP)
    courses["validacao_nome"] = courses["NO_CURSO"].map(
        lambda value: "Nome padrão" if normalize(value) == "HISTORIA" else "Nome especial - revisar"
    )
    courses["curso_fonte"] = "Inep/Censo da Educação Superior 2024"

    return courses.sort_values(["CO_IES", "NO_CURSO", "CO_CURSO"]).reset_index(drop=True)


def build_workbook(
    public_universities: pd.DataFrame,
    history_courses: pd.DataFrame,
    output: Path,
) -> None:
    counts = (
        history_courses.groupby("CO_IES")
        .agg(
            qtd_cursos_historia=("CO_CURSO", "nunique"),
            modalidades_historia=("modalidade", lambda s: ", ".join(sorted(set(s.dropna())))),
            graus_historia=("grau_academico", lambda s: ", ".join(sorted(set(s.dropna())))),
            matriculas_historia_2024=("QT_MAT", "sum"),
            vagas_historia_2024=("QT_VG_TOTAL", "sum"),
        )
        .reset_index()
    )

    universities = public_universities.merge(counts, on="CO_IES", how="left")
    universities["qtd_cursos_historia"] = universities["qtd_cursos_historia"].fillna(0).astype(int)
    universities["matriculas_historia_2024"] = universities["matriculas_historia_2024"].fillna(0).astype(int)
    universities["vagas_historia_2024"] = universities["vagas_historia_2024"].fillna(0).astype(int)
    universities["tem_curso_historia"] = universities["qtd_cursos_historia"].map(
        lambda value: "Sim" if value > 0 else "Não"
    )
    universities["email_secao_alunos"] = ""
    universities["email_secretaria_curso_historia"] = ""
    universities["email_departamento_historia"] = ""
    universities["email_chefe_departamento"] = ""
    universities["fonte_email"] = ""
    universities["url_fonte_email"] = ""
    universities["data_verificacao_email"] = ""
    universities["status_apuracao_email"] = universities["tem_curso_historia"].map(
        lambda value: "Pendente" if value == "Sim" else "Não aplicável"
    )
    universities["observacoes"] = ""

    universities_cols = [
        "CO_IES",
        "SG_IES",
        "NO_IES",
        "NO_REGIAO_IES",
        "SG_UF_IES",
        "NO_UF_IES",
        "NO_MUNICIPIO_IES",
        "categoria_administrativa",
        "tem_curso_historia",
        "qtd_cursos_historia",
        "modalidades_historia",
        "graus_historia",
        "matriculas_historia_2024",
        "vagas_historia_2024",
        "email_secao_alunos",
        "email_secretaria_curso_historia",
        "email_departamento_historia",
        "email_chefe_departamento",
        "fonte_email",
        "url_fonte_email",
        "data_verificacao_email",
        "status_apuracao_email",
        "observacoes",
        "endereco_ies",
        "NO_BAIRRO_IES",
        "NU_CEP_IES",
    ]
    universities = universities[universities_cols].rename(columns=humanize_columns)

    courses = history_courses.merge(
        public_universities[
            ["CO_IES", "SG_IES", "NO_IES", "SG_UF_IES", "NO_MUNICIPIO_IES"]
        ],
        on="CO_IES",
        how="left",
    )
    courses_cols = [
        "CO_IES",
        "SG_IES",
        "NO_IES",
        "SG_UF_IES",
        "NO_MUNICIPIO_IES",
        "CO_CURSO",
        "NO_CURSO",
        "NO_CINE_ROTULO",
        "grau_academico",
        "modalidade",
        "nivel_academico",
        "QT_CURSO",
        "QT_VG_TOTAL",
        "QT_MAT",
        "validacao_nome",
        "curso_fonte",
    ]
    courses = courses[courses_cols].rename(columns=humanize_columns)

    pending = universities[universities["Tem Curso Historia"] == "Sim"].copy()
    pending = pending[
        [
            "CO IES",
            "SG IES",
            "NO IES",
            "SG UF IES",
            "NO MUNICIPIO IES",
            "Categoria Administrativa",
            "Qtd Cursos Historia",
            "Email Secao Alunos",
            "Email Secretaria Curso Historia",
            "Email Departamento Historia",
            "Email Chefe Departamento",
            "Status Apuracao Email",
            "Url Fonte Email",
            "Observacoes",
        ]
    ]

    sprints = pd.DataFrame(
        [
            {
                "Sprint": "Sprint 0",
                "Objetivo": "Estruturar projeto, fontes oficiais, critérios e planilha-base.",
                "Escopo": "Baixar Censo/Inep 2024, listar universidades públicas e identificar cursos de História.",
                "Status": "Concluída",
                "Saída esperada": "Planilha com abas Universidades, Cursos_Historia, Pendencias_Email, Sprints e Fontes.",
            },
            {
                "Sprint": "Sprint 1",
                "Objetivo": "Apurar e-mails das universidades com História nas regiões Norte e Centro-Oeste.",
                "Escopo": "Verificar sites oficiais, páginas de departamento/curso, pró-reitorias e diretórios acadêmicos institucionais.",
                "Status": "Pendente",
                "Saída esperada": "E-mails preenchidos com URL da fonte e data de verificação.",
            },
            {
                "Sprint": "Sprint 2",
                "Objetivo": "Apurar e-mails das universidades com História no Nordeste.",
                "Escopo": "Mesmo método da Sprint 1, priorizando secretaria do curso/departamento; seção de alunos como fallback.",
                "Status": "Pendente",
                "Saída esperada": "E-mails preenchidos com URL da fonte e data de verificação.",
            },
            {
                "Sprint": "Sprint 3",
                "Objetivo": "Apurar e-mails das universidades com História no Sudeste.",
                "Escopo": "Mesmo método da Sprint 1.",
                "Status": "Pendente",
                "Saída esperada": "E-mails preenchidos com URL da fonte e data de verificação.",
            },
            {
                "Sprint": "Sprint 4",
                "Objetivo": "Apurar e-mails das universidades com História no Sul.",
                "Escopo": "Mesmo método da Sprint 1.",
                "Status": "Pendente",
                "Saída esperada": "E-mails preenchidos com URL da fonte e data de verificação.",
            },
            {
                "Sprint": "Sprint 5",
                "Objetivo": "Revisão final e controle de qualidade.",
                "Escopo": "Remover duplicidades, checar domínios oficiais, marcar ausentes e consolidar observações.",
                "Status": "Pendente",
                "Saída esperada": "Planilha final pronta para uso, com pendências justificadas.",
            },
        ]
    )

    sources = pd.DataFrame(
        [
            {
                "Fonte": "Inep - Microdados do Censo da Educação Superior 2024",
                "Uso": "Base oficial de IES, categoria administrativa e cursos.",
                "URL": "https://www.gov.br/inep/pt-br/acesso-a-informacao/dados-abertos/microdados/censo-da-educacao-superior",
                "Observação": "Arquivo ZIP: microdados_censo_da_educacao_superior_2024.zip.",
            },
            {
                "Fonte": "Sites oficiais das universidades",
                "Uso": "Apuração de e-mails de secretaria, seção de alunos, departamento ou chefia.",
                "URL": "",
                "Observação": "A preencher nas sprints de apuração; usar apenas fonte oficial ou institucional.",
            },
        ]
    )

    resumo = pd.DataFrame(
        [
            {"Métrica": "Data de geração", "Valor": date.today().isoformat()},
            {"Métrica": "Ano-base Censo/Inep", "Valor": 2024},
            {"Métrica": "Universidades públicas listadas", "Valor": len(universities)},
            {"Métrica": "Universidades públicas com curso de História", "Valor": int((universities["Tem Curso Historia"] == "Sim").sum())},
            {"Métrica": "Cursos de História encontrados", "Valor": len(courses)},
            {"Métrica": "Critério de universidade pública", "Valor": "TP_ORGANIZACAO_ACADEMICA=1 e TP_REDE=1"},
            {"Métrica": "Critério de curso de História", "Valor": "NO_CURSO contém História; linhas locais TP_DIMENSAO=1; História da Arte excluída"},
        ]
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        universities.to_excel(writer, sheet_name="Universidades", index=False)
        courses.to_excel(writer, sheet_name="Cursos_Historia", index=False)
        pending.to_excel(writer, sheet_name="Pendencias_Email", index=False)
        sprints.to_excel(writer, sheet_name="Sprints", index=False)
        sources.to_excel(writer, sheet_name="Fontes", index=False)

    format_workbook(output)


def humanize_columns(column: str) -> str:
    replacements = {
        "CO_IES": "CO IES",
        "SG_IES": "SG IES",
        "NO_IES": "NO IES",
        "NO_REGIAO_IES": "NO REGIAO IES",
        "SG_UF_IES": "SG UF IES",
        "NO_UF_IES": "NO UF IES",
        "NO_MUNICIPIO_IES": "NO MUNICIPIO IES",
        "CO_CURSO": "CO CURSO",
        "NO_CURSO": "NO CURSO",
        "NO_CINE_ROTULO": "NO CINE ROTULO",
        "QT_CURSO": "QT CURSO",
        "QT_VG_TOTAL": "QT VG TOTAL",
        "QT_MAT": "QT MAT",
        "NO_BAIRRO_IES": "NO BAIRRO IES",
        "NU_CEP_IES": "NU CEP IES",
    }
    if column in replacements:
        return replacements[column]
    return column.replace("_", " ").title()


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells[:250]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 52)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Gera planilha de universidades públicas brasileiras e cursos de História."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Caminho do XLSX de saída.",
    )
    args = parser.parse_args()
    require_inputs()
    universities = load_public_universities()
    courses = load_history_courses(set(universities["CO_IES"].astype(int)))
    build_workbook(universities, courses, args.output)
    print(f"Planilha gerada: {args.output}")
    print(f"Universidades públicas: {len(universities)}")
    print(f"Universidades com História: {courses['CO_IES'].nunique()}")
    print(f"Cursos de História: {len(courses)}")


if __name__ == "__main__":
    main()
